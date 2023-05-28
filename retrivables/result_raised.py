from tkinter import messagebox
from tkinter import ttk
from PIL import ImageTk, Image
import openpyxl as pxl
from openpyxl.styles import Alignment
import pickle
import tkinter as tk
import pandas as pd
import retrivables.dict_data as dict_data
import datetime

App_font = ("Sans-Serif", 13)
Body_font = ("Sans-Serif", 9)

# this tab not only lets a user upload a result chart but also file a trade
# which is a way for a user to export their trade into an excel spreadsheet
class result(tk.Frame):
    def __init__(self, parent, controller):
        import  pages.mainpage_raised as mainpage_raised ,pages.four_hour_raised as four_hour_raised, pages.one_hour_raised as one_hour_raised, pages.fifteen_min_raised as fifteen_min_raised
        tk.Frame.__init__(self, parent)

        self.test_img = ""
        self.resizingImg = ""
        self.new_img = ""
        self.check_img = ""

        self.sheet_val = tk.StringVar()

        self.conf_4h = None

        self.upload_pass = False
        self.cont_export = False
        

        self.tabs_bar = tk.Frame(self, bg="#DC7979")
        self.tabs_bar.grid(row=0, column=0, sticky="we")

        self.banner_lbl = tk.Label(self, text="Result",font=App_font ,width=120, height=5, bg="#BDBABA")
        self.banner_lbl.grid(row=1, column=0 )

        self.result_body = tk.Frame(self, bg="#FDFDFD", pady=20)
        self.result_body.grid(row=2, column=0, sticky="nsew", padx=(20,0))

        self.button21 = tk.Button(self.tabs_bar, text="Main", command=lambda: controller.show_frame(mainpage_raised.mainpage), borderwidth=1, relief="solid")
        self.button21.grid(row=0, column=1, padx=(5, 5), pady=(5, 5))

        self.button22 = tk.Button(self.tabs_bar, text="4 Hour", command=lambda: controller.show_frame(four_hour_raised.four_hour), borderwidth=1, relief="solid")
        self.button22.grid(row=0, column=2, padx=(5, 5), pady=(5, 5))

        self.button23 = tk.Button(self.tabs_bar, text="1 Hour", command=lambda: controller.show_frame(one_hour_raised.one_hour), borderwidth=1, relief="solid")
        self.button23.grid(row=0, column=3, padx=(5, 5), pady=(5, 5))

        self.button24 = tk.Button(self.tabs_bar, text="15 Min", command=lambda: controller.show_frame(fifteen_min_raised.fifteen_min), borderwidth=1, relief="solid")
        self.button24.grid(row=0, column=4, padx=(5, 5), pady=(5, 5))

        self.button25 = tk.Button(self.tabs_bar, text="Result", command=lambda: controller.show_frame(result), borderwidth=1, relief="solid")
        self.button25.grid(row=0, column=5, padx=(5, 5), pady=(5, 5))

        self.disp_result = tk.Label(self.result_body, width=135, height=33, image=self.test_img, text="This Chart Is Currently Empty \nClick 'Upload Result Chart' To Add One" ,relief="solid", borderwidth=1)
        self.disp_result.grid(rowspan=30, columnspan=5)

        self.result_button = ttk.Button(self.result_body, text="       Add\nResult Chart", command=lambda: result.add_result(self))
        self.result_button.grid(row=14, column=5, padx=(20, 0))

        self.upload_result = ttk.Button(self.result_body, text="    Upload\nResult Chart", state="disabled" ,command=lambda: result.upload_result(self))
        self.upload_result.grid(row=15, column=5, sticky="E")

        self.file_trade = ttk.Button(self.result_body, text="  File\nTrade", command=lambda:result.file_trade(self, controller))
        self.file_trade.grid(row=25, column=5, padx=(20,0))

        self.txt_result_lbl = ttk.Label(self.result_body, text="Loss/Gain")
        self.txt_result_lbl.grid(row=1, column=5)

        self.result_lbl = tk.Label(self.result_body, text=dict_data.get_filedict_data.get("loss-gain"), relief="sunken", bg="white" ,borderwidth=1, width=9)
        self.result_lbl.grid(row=2, column=5, padx=(14,0))

        # enables or disables chart upload button depending on if trade is open or closed
        def start_result_tab(self):
            if dict_data.is_trade_placed is True:
                self.upload_result.config(state="normal")
            else:
                self.upload_result.config(state="disabled")

        # places added chart on a label if it was uploaded in the past
        def get_img(self):
            try:
                dict_data.filetrade = open("ft_data", "rb")
                dict_data.get_filedict_data = pickle.load(dict_data.filetrade)

                self.test_img = Image.open(dict_data.get_filedict_data.get("result-chart"))
                self.resizingImg = self.test_img.resize((950, 490), Image.ANTIALIAS)
                self.new_img = ImageTk.PhotoImage(self.resizingImg)
                self.disp_result.config(width=950, height=490 ,image=self.new_img, text="")
            except Exception:
                pass

        start_result_tab(self)
        get_img(self)
    
    # adds a result chart selected by a user
    def add_result(self):
        start_vars = open("startup_vars", "rb")
        dict_data.get_dict_data = pickle.load(start_vars)
        if dict_data.st_values_holder.get("is_placed") == False and dict_data.st_values_holder.get("is_trade_filed") == False:
            try:
                filename = tk.filedialog.askopenfile(initialdir="Desktop", title="Select File", filetypes=(("PNG Files", "*.png"),("JPG Files", "*.jpg")))
                dict_data.placetrade_values_holder["result-chart"] = filename.name
                self.test_img = Image.open(dict_data.placetrade_values_holder.get("result-chart"))
                self.resizingImg = self.test_img.resize((950, 490), Image.ANTIALIAS)
                self.new_img = ImageTk.PhotoImage(self.resizingImg)
                self.disp_result.config(width=950, height=490 ,image=self.new_img)
                self.upload_result.config(state="normal")
            except Exception:
                messagebox.showinfo("Error", "Failed to load the image")
        elif dict_data.st_values_holder.get("is_placed") == False and dict_data.st_values_holder.get("is_trade_filed") == True:
            messagebox.showinfo("Trade", "You have already filed a trade")
        else:
            messagebox.showinfo("Trade", "You haven't closed your recent trade")

    # uploads a result chart selected by a user
    def upload_result(self):
        start_vars = open("startup_vars", "rb")
        dict_data.get_dict_data = pickle.load(start_vars)

        dict_data.filetrade = open("ft_data", "rb")
        dict_data.get_filedict_data = pickle.load(dict_data.filetrade)

        if dict_data.placetrade_values_holder["result-chart"] != "" and dict_data.st_values_holder.get("is_placed") == True:
            messagebox.showinfo("Trade Error", "You Haven't Closed Your Trade Yet")
            self.cont_export = False
        elif dict_data.st_values_holder.get("is_trade_filed") == True:
            messagebox.showinfo("Trade Filed", "You Have Recently Filed A Trade")
            self.cont_export = False
        elif dict_data.placetrade_values_holder["result-chart"] != "" and dict_data.st_values_holder.get("is_placed") == False:
            messagebox.showinfo("Updated", "Chart Was Successfully Uploaded")
            self.upload_pass = True
            self.cont_export = True
            dict_data.placetrade_values_to_startup()
            dict_data.fileload_values_to_filedata()
            dict_data.ft_data["result-chart"] = dict_data.st_values_holder.get("result-chart")

            dict_data.filetrade = open("ft_data", "wb")
            pickle.dump(dict_data.ft_data, dict_data.filetrade)
            dict_data.filetrade.close()

            self.result_lbl.config(text=dict_data.ft_data.get("loss-gain"))

        else:
            messagebox.showinfo("No Chart", "You haven't selected any chart to upload")
            self.cont_export = False

    def validate_uploaded_chart(self):
        if dict_data.ft_data.get("result-chart") == "" and self.upload_pass == False:
            ask_if_uploaded = messagebox.askyesno("Un-uploaded File", "Result Chart Hasn't Been Uploaded\n    Do You Wish To Upload It ?")
            if ask_if_uploaded == True:
                result.upload_result(self)
            elif ask_if_uploaded == False:
                self.cont_export = True

    def remove_chart(self):
        self.disp_result.config(image="", width=135, height=33, text="This Chart Is Currently Empty \nClick 'Upload Result Chart' To Add One")

    def update_lossgain(self):
        dict_data.filetrade = open("ft_data", "rb")
        loss_gain = pickle.load(dict_data.filetrade)
        loss_gain = loss_gain.get("loss-gain")
        dict_data.filetrade.close()
        self.result_lbl.config(text=str(loss_gain)) 
        
    # operation which will export a trade into an excel file
    # user will be presented with a choice of either uploading data or not
    # data in dictionaries used by a user will be reset upon completion
    def file_trade(self, controller):
        import pages.mainpage_raised as mainpage_raised
        dict_data.filetrade = open("ft_data", "rb")
        dict_data.get_filedict_data = pickle.load(dict_data.filetrade)

        start_vars = open("startup_vars", "rb")
        dict_data.get_dict_data = pickle.load(start_vars)
    
    # validation checks ensuring user is aware that their result chart hasn't been uploaded
        result.validate_uploaded_chart(self)
        if self.cont_export == True:
            if dict_data.st_values_holder.get("is_placed") == False and dict_data.st_values_holder.get("is_trade_filed") == False:
                q2_answer = messagebox.askyesno("File", "Do you wish to file a trade ?\n\nNo - close a trade without filing\nYes - close and file a trade")
                if q2_answer is True:

                    result.check_if_exists(self)

                    dict_data.filetrade = open("ft_data", "rb")
                    dict_data.get_filedict_data = pickle.load(dict_data.filetrade)

                    # load recently picked trade with it's parameters/values
                    dict_data.startup_load_values_to_startup()
                    dict_data.st_values_holder["is_trade_filed"] = True
                    dict_data.st_values_holder["4-hour"] = ""
                    dict_data.st_values_holder["1-hour"] = ""
                    dict_data.st_values_holder["15-min"] = ""
                    dict_data.st_values_holder["result-chart"] = ""

                    dict_data.placetrade_values_holder["4-hour"] = ""
                    dict_data.placetrade_values_holder["1-hour"] = ""
                    dict_data.placetrade_values_holder["15-min"] = ""
                    dict_data.placetrade_values_holder["result-chart"] = ""

                    # pickele-save our start-up values for later use (incase the app closes) 
                    start_up = open("startup_vars", "wb")
                    pickle.dump(dict_data.st_values_holder, start_up)
                    start_up.close()

                    dict_data.export_data()
                    result.export(self)
                    
                    dict_data.reset_filedata_dict()
                    # move back and display mainpage
                    controller.show_frame(mainpage_raised.mainpage)
                    
                elif q2_answer is False:

                    # load recently picked trade with it's parameters/values
                    dict_data.startup_load_values_to_startup()
                    dict_data.st_values_holder["is_trade_filed"] = True
                    dict_data.st_values_holder["4-hour"] = ""
                    dict_data.st_values_holder["1-hour"] = ""
                    dict_data.st_values_holder["15-min"] = ""
                    dict_data.st_values_holder["result-chart"] = ""

                    dict_data.placetrade_values_holder["4-hour"] = ""
                    dict_data.placetrade_values_holder["1-hour"] = ""
                    dict_data.placetrade_values_holder["15-min"] = ""
                    dict_data.placetrade_values_holder["result-chart"] = ""

                    # pickle-save our start-up values for later use (incase the app closes)
                    start_up = open("startup_vars", "wb")
                    pickle.dump(dict_data.st_values_holder, start_up)
                    start_up.close()

                    dict_data.reset_filedata_dict()
                    controller.show_frame(mainpage_raised.mainpage)

                    self.result_lbl.config(text=0)
            elif dict_data.st_values_holder.get("is_placed") is False and dict_data.st_values_holder.get("is_trade_filed") is True:
                messagebox.showinfo("Trade", "You have already filed this trade")
            elif dict_data.st_values_holder.get("is_placed") is True and dict_data.st_values_holder.get("is_trade_filed") is False:
                messagebox.showinfo("Trade", "You're currently in a trade")
            elif dict_data.st_values_holder.get("is_placed") is True and dict_data.st_values_holder.get("is_trade_filed") is True:
                messagebox.showinfo("Trade", "          You already filed a trade\n Close your current trade to file again")
        else:
            pass

    # check if the charts exist, if one or more do not exist, the chart that is missing will be set to NULL
    def check_if_exists(self):
        dict_data.filetrade = open("ft_data", "rb")
        dict_data.get_filedict_data = pickle.load(dict_data.filetrade)
        try:
            if dict_data.get_filedict_data.get("4-hour") != "":
                try:
                    self.check_img = Image.open(dict_data.get_filedict_data.get("4-hour"))
                except Exception:
                    dict_data.fileload_values_to_filedata()
                    dict_data.ft_data["4-hour"] = ""
                    
                    dict_data.filetrade = open("ft_data", "wb")
                    pickle.dump(dict_data.ft_data, dict_data.filetrade)
                    dict_data.filetrade.close()
            if dict_data.get_filedict_data.get("1-hour") != "":
                try:
                    self.check_img = Image.open(dict_data.get_filedict_data.get("1-hour"))
                except Exception:
                    dict_data.fileload_values_to_filedata()
                    dict_data.ft_data["1-hour"] = ""
                    
                    dict_data.filetrade = open("ft_data", "wb")
                    pickle.dump(dict_data.ft_data, dict_data.filetrade)
                    dict_data.filetrade.close()
            if dict_data.get_filedict_data.get("15-min") != "":
                try:
                    self.check_img = Image.open(dict_data.get_filedict_data.get("15-min"))
                except Exception:
                    dict_data.fileload_values_to_filedata()
                    dict_data.ft_data["15-min"] = ""
                    
                    dict_data.filetrade = open("ft_data", "wb")
                    pickle.dump(dict_data.ft_data, dict_data.filetrade)
                    dict_data.filetrade.close()
            if dict_data.get_filedict_data.get("result-chart") != "":
                try:
                    self.check_img = Image.open(dict_data.get_filedict_data.get("result-chart"))
                except Exception:
                    dict_data.fileload_values_to_filedata()
                    dict_data.ft_data["result-chart"] = ""
                    
                    dict_data.filetrade = open("ft_data", "wb")
                    pickle.dump(dict_data.ft_data, dict_data.filetrade)
                    dict_data.filetrade.close()
        except Exception:
            pass

    
        
    # loads in the data from the dict_data.ready_for_export dictionary into excel workbook and adds it to the pandas dataframe
    # if the workbook doesn't exist, it creates new one with a current date and number representing worksheet number of "1"
    # else if the workbook does exist, it creates new worksheet with current date and number representing the total of existing worhseets within the workbook
    # charts & cells are resized to fit in the resized excel cell, if any chart is missing, they will end-up not beeing uploaded
    def export(self):
        with open("ft_data", "rb") as filetrade:
            dict_data.get_filedict_data = pickle.load(filetrade)

        current_date_time = datetime.datetime.now()
        formatted_year = str(current_date_time.year)[2:]
        current_date = str(current_date_time.day) + "-" + str(current_date_time.month) + "-" + formatted_year
        filename = 'ForexTrades.xlsx'

    # checks if the sheetname with a current date already exisits within the workbook
    # if sheet name does exist, new name is assigned to it with a current date and number count representing the ith trade placed on that day
    # trades filed on a new day will have number starting from '1' assigned to their trade sheet name 
        try:
            book = pxl.load_workbook(filename)
            i = 1
            sheet_name = str(current_date) + " - " + str(i)
            while sheet_name in book.sheetnames:
                i += 1
                sheet_name = str(current_date) + " - " + str(i)
                

            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
            self.write_to_excel(filename, writer, sheet_name)
        except FileNotFoundError:
            sheet_name = str(current_date) + " - " + str(1)
            messagebox.showinfo("No Workbook", "Failed To Find Workbook \n    Creating New One")
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            self.write_to_excel(filename, writer, sheet_name)
    

    def write_to_excel(self, filename, writer, sheet_name):
        data = {'Pair': [dict_data.ready_for_export.get("Pair")], 'Lot-size': [dict_data.ready_for_export.get("Lot-size")], 'Buy-sell': [dict_data.ready_for_export.get("Buy-sell")],
                'Open': [dict_data.ready_for_export.get("Open")], 'Close': [dict_data.ready_for_export.get("Close")], 'Balance': [("£" + str(dict_data.ready_for_export.get("Balance")))],
                'Pos-size': [dict_data.ready_for_export.get("Pos-size")], 'Loss-gain': [("£" + str(dict_data.ready_for_export.get("Loss-gain")))], '4-hour': [dict_data.ready_for_export.get("4-hour")],
                '1-hour': [dict_data.ready_for_export.get("1-hour")], '15-min': [dict_data.ready_for_export.get("15-min")], 'Result': [dict_data.ready_for_export.get("Result-chart")]}
        
        df = pd.DataFrame(data)
        df.to_excel(writer, sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        try:
            four_h = dict_data.ready_for_export.get("4-hour")
            one_h = dict_data.ready_for_export.get("1-hour")
            fifteen_min = dict_data.ready_for_export.get("15-min")
            result_chart = dict_data.ready_for_export.get("Result-chart")
        except Exception:
            pass

        self.add_image(worksheet, "I", four_h)
        self.add_image(worksheet, "J", one_h)
        self.add_image(worksheet, "K", fifteen_min)
        self.add_image(worksheet, "L", result_chart)

    # Resizes cells to accomodate space for user typed notes and confuluences for each filed trade
        worksheet.merge_cells("B7:G12")
        cell = worksheet.cell(row=7, column=2)
        cell.value = "Notes:"
        cell.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)

        worksheet.merge_cells("B15:G18")
        cell2 = worksheet.cell(row=15, column=2)
        cell2.value = "Confluences:"
        cell2.alignment = Alignment(horizontal="left", vertical="top", wrapText=True)
        
        writer.close()
        self.delete_sheet_columns(filename, sheet_name)


    # Get's rid of all unused columns in excel spreadsheet
    def delete_sheet_columns(self, filename, sheet_name):
        book = pxl.load_workbook(filename)
        sheet = book[sheet_name]
        sheet.column_dimensions.group(start='M', end='XFD', hidden=True)
        book.save(filename)


    # Resizes cells and adds image to a cell if added
    def add_image(self, worksheet, column, file):
        try:
            worksheet.column_dimensions[column].width = 150
            img_display = Image.open(file)
            img = pxl.drawing.image.Image(img_display)
            img.width = 1050
            img.height = 600
            img.anchor = f"{column}2"
            worksheet.add_image(img)
        except Exception:
            pass
        


        

        
            







        